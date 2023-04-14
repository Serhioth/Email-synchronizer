import imaplib
import logging
import os
import random
import re
import string
from http import HTTPStatus
from logging.handlers import RotatingFileHandler

import requests
from dotenv import load_dotenv
from emails.constants import (DOMAINS, IMAP_SERVER, SYNC_INTERVAL,
                              URL_NEW_MAIL, URL_NEW_SYNC)
from emails.exceptions import (DomainError, FIOError, IMAPError,
                               ParseEmailError, RequestsError)
from imapclient.imapclient import IMAPClient
from openpyxl import load_workbook

load_dotenv()

API_KEY = os.environ['API_KEY']
logic_logger = logging.getLogger(name=__name__)
logic_logger.setLevel('INFO')

handler = RotatingFileHandler(
    'main_log.log',
    maxBytes=50000000,
    backupCount=5,
    encoding='utf-8'
)
formatter = logging.Formatter(
    '%(asctime)s, %(levelname)s, %(name)s, '
    '%(funcName)s, %(levelno)s, %(message)s'
)

handler.setFormatter(formatter)
logic_logger.addHandler(handler)


def generate_password() -> string:
    """Generate password. """
    logic_logger.debug('Генерируется пароль')
    chars = string.ascii_letters + string.digits
    password = ''.join(random.choice(chars) for i in range(10))
    return password


def check_user_imap(email: str, password: str) -> bool:
    """Check user can connect to IMAP server. """
    logic_logger.debug('мы дошли до АЙМАПА')
    server = IMAPClient(IMAP_SERVER, use_uid=True, ssl=True)
    try:
        server.login(email, password)
        logic_logger.info(f'Выполнено успешное подключение к {IMAP_SERVER}')
        return True
    except imaplib.IMAP4.error as e:
        logic_logger.error(IMAPError(e))
        return False


def create_new_account(
    login: str,
    password: str,
    domain: str,
    email_yandex: str,
    password_yandex: str,
    fio: str
):
    """Crate new account for user. """
    logic_logger.debug('Создаём аккаунт')

    headers = {
            'accept': 'application/json',
            'X-API-Key': API_KEY,
            'Content-Type': 'application/json'
        }

    body = {
            "active": "1",
            "domain": domain,
            "local_part": login,
            "name": fio,
            "password": password,
            "password2": password,
            'quota': '10240',
            'force_pw_update': '0',
            "tls_enforce_in": "1",
            "tls_enforce_out": "1",
        }

    body2 = {
        "username": email_yandex,
        "host1": "imap.yandex.ru",
        "port1": "993",
        "user1": email_yandex,
        "password1": password_yandex,
        "enc1": "TLS",
        "mins_interval": SYNC_INTERVAL,
        "subfolder2": "/",
        "maxage": "0",
        "maxbytespersecond": "0",
        "timeout1": "100",
        "timeout2": "100",
        "exclude": "(?i)spam|(?i)junk",
        "delete2duplicates": "1",
        "delete1": "0",
        "delete2": "0",
        "automap": "1",
        "skipcrossduplicates": "0",
        "subscribeall": "0",
        "active": "1"
    }
    logic_logger.debug('Вот тут реквест')
    try:
        response_new_email = requests.post(
            url=URL_NEW_MAIL,
            headers=headers,
            json=body
        )
        logic_logger.info(response_new_email.text)
    except Exception as e:
        raise e
    if response_new_email.status_code == HTTPStatus.OK:
        new_mail_status = response_new_email.json()[0]['msg'][0]
        logic_logger.info(new_mail_status)
        if new_mail_status != 'mailbox_added':
            return new_mail_status
        try:
            response_sync = requests.post(
                url=URL_NEW_SYNC,
                headers=headers,
                json=body2
            )
        except Exception as e:
            logic_logger.error(RequestsError(e))
            raise RequestsError(e)
        sync_status = response_sync.json()[0]['msg'][0]
        logic_logger.info(sync_status)
        return new_mail_status


def parse_email(email: str):
    """Get login and domain from email. """
    logic_logger.debug('Разбираем email')
    pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    if re.match(pattern, email):
        login, domain = email.split('@')
        if domain not in DOMAINS:
            logic_logger.error(DomainError(domain))
            raise DomainError(domain)
        return login, domain
    logic_logger.error(ParseEmailError(email))
    raise ParseEmailError(email)


def get_fio(login: str, domain: str):
    """Get user's FIO by login. """
    logic_logger.debug('Ищем ФИО')
    if domain == 'mokeev.ru':
        xl_file = load_workbook(
            # Replace dummy with name of your excel-file
            'emails/xl-files/dummy-domain-data.xlsx',
            data_only=True
        )
        xl_file.active = 0
        work_list = xl_file.active
        fio_col = work_list['A']  # Replace with index of your column
        login_col = work_list['B']  # Replace with index of your column
        login_fio = dict()
        key = login

        for login, fio in zip(login_col, fio_col):
            login_fio[login.value] = fio.value
        if key in login_fio:
            return login_fio[key]
        else:
            logic_logger.error(FIOError(key))
            return key
    xl_file = load_workbook(
        # Replace dummy with name of your excel-file
        'emails/xl-files/dummy-sec-domain-data.xlsx',
        data_only=True
    )
    xl_file.active = 0
    work_list = xl_file.active
    fio_col = work_list['A']  # Replace with index of your column
    login_col = work_list['B']  # Replace with index of your column
    login_fio = dict()
    key = login

    for login, fio in zip(login_col, fio_col):
        login_fio[login.value] = fio.value
    if key in login_fio:
        return login_fio[key]
    else:
        logic_logger.error(FIOError(key))
        return key
